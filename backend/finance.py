"""
AelfLab Finance — Backend API
Database + endpoints untuk pencatatan keuangan pribadi

Catatan semantik saldo (dipertahankan agar konsisten dengan data lama):
  income, saving -> menambah saldo akun
  expense        -> mengurangi saldo akun
  transfer       -> mengurangi akun asal, menambah akun tujuan
  invest         -> TIDAK mengubah saldo akun
Perubahan aturan di atas akan mengubah saldo historis, jadi jangan diubah
tanpa migrasi data yang disengaja.
"""
import calendar
import io
import json
import re
import sqlite3
import subprocess
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import List, Optional

import openpyxl
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel

HUB_DIR = Path(__file__).resolve().parent.parent
DB_PATH = HUB_DIR / "data" / "finance.db"

router = APIRouter(prefix="/api/finance", tags=["finance"])

TX_TYPES = ("income", "expense", "transfer", "invest", "saving")
FREQUENCIES = ("daily", "weekly", "monthly")

# Tipe yang menaikkan / menurunkan saldo akun utama
_PLUS_TYPES = ("income", "saving")
_MINUS_TYPES = ("expense",)


# ── Database ──────────────────────────────────────

def get_db():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def _columns(conn, table: str) -> List[str]:
    return [r["name"] for r in conn.execute(f"PRAGMA table_info({table})")]


def _migrate(conn):
    """Tambah kolom baru pada database yang sudah ada.

    Dipanggil setiap start; aman dijalankan berulang.
    """
    if "last_run" not in _columns(conn, "recurring"):
        conn.execute("ALTER TABLE recurring ADD COLUMN last_run TEXT")
    if "recurring_id" not in _columns(conn, "transactions"):
        conn.execute("ALTER TABLE transactions ADD COLUMN recurring_id INTEGER")
    if "installment_id" not in _columns(conn, "transactions"):
        conn.execute("ALTER TABLE transactions ADD COLUMN installment_id INTEGER")
    conn.commit()


def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            type TEXT NOT NULL DEFAULT 'cash',
            balance REAL NOT NULL DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            icon TEXT DEFAULT '📦',
            type TEXT NOT NULL DEFAULT 'expense'
        );

        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL DEFAULT (date('now','localtime')),
            time TEXT DEFAULT (time('now','localtime')),
            type TEXT NOT NULL CHECK(type IN ('income','expense','transfer','invest','saving')),
            amount REAL NOT NULL,
            category_id INTEGER REFERENCES categories(id),
            account_id INTEGER REFERENCES accounts(id),
            account_to_id INTEGER REFERENCES accounts(id),
            note TEXT,
            merchant TEXT,
            recurring_id INTEGER,
            installment_id INTEGER,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS budgets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category_id INTEGER REFERENCES categories(id),
            month TEXT NOT NULL,
            amount REAL NOT NULL,
            spent REAL DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS recurring (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            type TEXT NOT NULL DEFAULT 'expense',
            amount REAL NOT NULL,
            category_id INTEGER REFERENCES categories(id),
            account_id INTEGER REFERENCES accounts(id),
            frequency TEXT NOT NULL DEFAULT 'monthly',
            day INTEGER NOT NULL DEFAULT 1,
            next_date TEXT,
            last_run TEXT,
            active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS installments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            total_amount REAL NOT NULL,
            total_installments INTEGER NOT NULL,
            paid_installments INTEGER DEFAULT 0,
            monthly_amount REAL NOT NULL,
            category_id INTEGER REFERENCES categories(id),
            account_id INTEGER REFERENCES accounts(id),
            start_date TEXT NOT NULL,
            end_date TEXT,
            active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS investments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            type TEXT NOT NULL CHECK(type IN ('crypto','gold','stock','mutual_fund','forex','deposit','other')),
            amount REAL NOT NULL,
            quantity REAL,
            buy_price REAL,
            current_price REAL,
            currency TEXT DEFAULT 'IDR',
            account_id INTEGER REFERENCES accounts(id),
            maturity_date TEXT,
            return_rate REAL,
            note TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE INDEX IF NOT EXISTS idx_tx_date ON transactions(date);
        CREATE INDEX IF NOT EXISTS idx_tx_type ON transactions(type);
        CREATE INDEX IF NOT EXISTS idx_tx_cat ON transactions(category_id);
        CREATE INDEX IF NOT EXISTS idx_tx_acct ON transactions(account_id);
    """)
    conn.commit()
    _migrate(conn)

    cats = [
        ('Makanan', '🍜', 'expense'), ('Minuman', '☕', 'expense'),
        ('Transport', '🚗', 'expense'), ('Belanja', '🛒', 'expense'),
        ('Hiburan', '🎮', 'expense'), ('Kesehatan', '💊', 'expense'),
        ('Tagihan', '📋', 'expense'), ('Kos/Sewa', '🏠', 'expense'),
        ('Gaji', '💰', 'income'), ('Bonus', '🎁', 'income'),
        ('Tabungan', '🏦', 'saving'), ('Investasi', '📈', 'invest'),
    ]
    for name, icon, typ in cats:
        conn.execute("INSERT OR IGNORE INTO categories(name, icon, type) VALUES (?,?,?)", (name, icon, typ))

    for acct in [('Tunai', 'cash'), ('BCA', 'bank'), ('Jenius', 'bank'), ('GoPay', 'ewallet')]:
        conn.execute("INSERT OR IGNORE INTO accounts(name, type) VALUES (?,?)", acct)

    conn.commit()
    conn.close()

init_db()


# ── Models ────────────────────────────────────────

class TransactionIn(BaseModel):
    type: str
    amount: float
    category: Optional[str] = None
    account: Optional[str] = None
    account_to: Optional[str] = None
    note: Optional[str] = None
    merchant: Optional[str] = None
    date: Optional[str] = None

class TransactionPatch(BaseModel):
    type: Optional[str] = None
    amount: Optional[float] = None
    category: Optional[str] = None
    account: Optional[str] = None
    account_to: Optional[str] = None
    note: Optional[str] = None
    merchant: Optional[str] = None
    date: Optional[str] = None

class AccountIn(BaseModel):
    name: str
    type: str = 'cash'

class BudgetIn(BaseModel):
    category: str
    month: str
    amount: float

class RecurringIn(BaseModel):
    name: str
    type: str = 'expense'
    amount: float
    category: Optional[str] = None
    account: Optional[str] = None
    frequency: str = 'monthly'
    day: int = 1

class RecurringPatch(BaseModel):
    name: Optional[str] = None
    amount: Optional[float] = None
    day: Optional[int] = None
    frequency: Optional[str] = None
    active: Optional[bool] = None

class InstallmentIn(BaseModel):
    name: str
    total_amount: float
    total_installments: int
    monthly_amount: float
    category: Optional[str] = None
    account: Optional[str] = None
    start_date: Optional[str] = None

class InvestmentIn(BaseModel):
    name: str
    type: str
    amount: float
    quantity: Optional[float] = None
    buy_price: Optional[float] = None
    account: Optional[str] = None
    maturity_date: Optional[str] = None
    return_rate: Optional[float] = None
    note: Optional[str] = None


# ── Helpers ───────────────────────────────────────

def get_or_create(conn, table, name_field, name, extra=None):
    row = conn.execute(f"SELECT id FROM {table} WHERE {name_field}=?", (name,)).fetchone()
    if row:
        return row['id']
    if extra:
        conn.execute(f"INSERT INTO {table}({name_field}, {extra[0]}) VALUES (?,?)", (name, extra[1]))
    else:
        conn.execute(f"INSERT INTO {table}({name_field}) VALUES (?)", (name,))
    conn.commit()
    return conn.execute(f"SELECT id FROM {table} WHERE {name_field}=?", (name,)).fetchone()['id']


def _validate_type(t: str):
    if t not in TX_TYPES:
        raise HTTPException(400, f"Tipe tidak valid: {t}. Pilihan: {', '.join(TX_TYPES)}")


def _validate_amount(a: float):
    if a is None or a <= 0:
        raise HTTPException(400, "Jumlah harus lebih besar dari 0")


def _validate_date(s: Optional[str]) -> Optional[str]:
    if not s:
        return None
    try:
        return date.fromisoformat(s[:10]).isoformat()
    except ValueError:
        raise HTTPException(400, f"Format tanggal harus YYYY-MM-DD, dapat: {s}")


def _validate_month(s: str) -> str:
    if not re.fullmatch(r"\d{4}-\d{2}", s or ""):
        raise HTTPException(400, f"Format bulan harus YYYY-MM, dapat: {s}")
    return s


def _apply_balance(conn, tx, sign: int):
    """Terapkan (sign=+1) atau batalkan (sign=-1) efek transaksi ke saldo akun.

    Satu fungsi untuk create/update/delete supaya saldo tidak pernah lepas
    sinkron dari daftar transaksi.
    """
    typ = tx["type"]
    amt = (tx["amount"] or 0) * sign
    acct = tx["account_id"]
    acct_to = tx["account_to_id"]

    if typ in _PLUS_TYPES and acct:
        conn.execute("UPDATE accounts SET balance = balance + ? WHERE id=?", (amt, acct))
    elif typ in _MINUS_TYPES and acct:
        conn.execute("UPDATE accounts SET balance = balance - ? WHERE id=?", (amt, acct))
    elif typ == 'transfer':
        if acct:
            conn.execute("UPDATE accounts SET balance = balance - ? WHERE id=?", (amt, acct))
        if acct_to:
            conn.execute("UPDATE accounts SET balance = balance + ? WHERE id=?", (amt, acct_to))
    # 'invest' sengaja tidak mengubah saldo — lihat catatan di docstring modul


def _sync_budget_spent(conn, month: str, category_id: Optional[int] = None):
    """Hitung ulang kolom spent dari transaksi sebenarnya.

    Versi lama menaikkan spent secara inkremental dan hanya kalau baris
    budget sudah ada, sehingga mudah lepas sinkron. Sekarang selalu
    diturunkan dari data transaksi.
    """
    sql = "SELECT id, category_id FROM budgets WHERE month=?"
    params = [month]
    if category_id is not None:
        sql += " AND category_id=?"
        params.append(category_id)
    for b in conn.execute(sql, params).fetchall():
        spent = conn.execute(
            "SELECT COALESCE(SUM(amount),0) FROM transactions "
            "WHERE type='expense' AND category_id=? AND date LIKE ?",
            (b["category_id"], month + '%')
        ).fetchone()[0]
        conn.execute("UPDATE budgets SET spent=? WHERE id=?", (spent, b["id"]))


def _days_in_month(y: int, m: int) -> int:
    return calendar.monthrange(y, m)[1]


def _month_shift(d: date, months: int, day: int) -> date:
    """Geser bulan sambil menjaga tanggal target, di-clamp ke akhir bulan.

    Tanggal 31 pada Februari menjadi 28/29, bukan error.
    """
    total = (d.year * 12 + (d.month - 1)) + months
    y, m = divmod(total, 12)
    m += 1
    return date(y, m, min(day, _days_in_month(y, m)))


def _first_due(freq: str, day: int, today: Optional[date] = None) -> date:
    """Jatuh tempo berikutnya, termasuk hari ini kalau memang pas."""
    today = today or date.today()
    if freq == 'daily':
        return today
    if freq == 'weekly':
        target = max(0, min(int(day), 6))          # 0=Senin
        delta = (target - today.weekday()) % 7
        return today + timedelta(days=delta)
    day = max(1, min(int(day), 31))
    candidate = date(today.year, today.month, min(day, _days_in_month(today.year, today.month)))
    return candidate if candidate >= today else _month_shift(today, 1, day)


def _advance(freq: str, day: int, current: date) -> date:
    if freq == 'daily':
        return current + timedelta(days=1)
    if freq == 'weekly':
        return current + timedelta(days=7)
    return _month_shift(current, 1, max(1, min(int(day), 31)))


def fetch_crypto_price(coin_id: str = "bitcoin") -> Optional[float]:
    """Get current BTC/ETH price in IDR from CoinGecko."""
    try:
        r = subprocess.run(
            ["curl", "-sk", f"https://api.coingecko.com/api/v3/simple/price?ids={coin_id}&vs_currencies=idr"],
            capture_output=True, text=True, timeout=10
        )
        data = json.loads(r.stdout)
        return data.get(coin_id, {}).get("idr")
    except Exception:
        return None


def fetch_gold_price() -> Optional[float]:
    """Get current gold price per gram in IDR."""
    try:
        r = subprocess.run(
            ["curl", "-sk", "https://harga-emas.org/"],
            capture_output=True, text=True, timeout=10
        )
        for line in r.stdout.splitlines():
            if "1 gram" in line and "Rp" in line:
                m = re.search(r'Rp[.\s]*([\d,.]+)', line)
                if m:
                    return float(m.group(1).replace('.', '').replace(',', ''))
    except Exception:
        return None
    return None


TX_SELECT = """
    SELECT t.*, c.name as cat_name, c.icon, a.name as acct_name, a2.name as acct_to_name
    FROM transactions t
    LEFT JOIN categories c ON t.category_id = c.id
    LEFT JOIN accounts a ON t.account_id = a.id
    LEFT JOIN accounts a2 ON t.account_to_id = a2.id
"""


# ── Meta ──────────────────────────────────────────

@router.get("/init")
def api_init():
    """Re-seed database (safe to call anytime)."""
    init_db()
    return {"status": "ok"}


@router.get("/categories")
def list_categories():
    conn = get_db()
    rows = conn.execute("SELECT * FROM categories ORDER BY type, name").fetchall()
    conn.close()
    return {"categories": [dict(r) for r in rows]}


@router.get("/accounts")
def list_accounts():
    conn = get_db()
    rows = conn.execute("SELECT * FROM accounts ORDER BY name").fetchall()
    conn.close()
    return {"accounts": [dict(r) for r in rows]}


@router.post("/accounts")
def add_account(a: AccountIn):
    if not a.name.strip():
        raise HTTPException(400, "Nama akun tidak boleh kosong")
    conn = get_db()
    try:
        conn.execute("INSERT INTO accounts(name, type) VALUES (?,?)", (a.name.strip(), a.type))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        raise HTTPException(409, f"Akun '{a.name}' sudah ada")
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass
    return {"status": "ok", "account": {"name": a.name.strip(), "type": a.type, "balance": 0}}


@router.post("/recalc")
def recalc_balances():
    """Hitung ulang seluruh saldo akun dari daftar transaksi.

    Jaring pengaman kalau saldo pernah lepas sinkron (mis. akibat edit
    manual di DB).
    """
    conn = get_db()
    before = {r["name"]: r["balance"] for r in conn.execute("SELECT name, balance FROM accounts")}
    conn.execute("UPDATE accounts SET balance = 0")
    for tx in conn.execute("SELECT * FROM transactions").fetchall():
        _apply_balance(conn, tx, +1)
    conn.commit()
    after = {r["name"]: r["balance"] for r in conn.execute("SELECT name, balance FROM accounts")}
    for month in [r["m"] for r in conn.execute("SELECT DISTINCT month AS m FROM budgets")]:
        _sync_budget_spent(conn, month)
    conn.commit()
    conn.close()
    changed = {k: {"before": before.get(k), "after": v}
               for k, v in after.items() if abs((before.get(k) or 0) - v) > 0.005}
    return {"status": "ok", "accounts": after, "changed": changed}


# ── Transactions ──────────────────────────────────

@router.post("/transactions")
def add_transaction(t: TransactionIn):
    _validate_type(t.type)
    _validate_amount(t.amount)
    tx_date = _validate_date(t.date) or date.today().isoformat()

    conn = get_db()
    cat_id = None
    if t.category:
        cat_id = get_or_create(conn, 'categories', 'name', t.category,
                               ('type', t.type if t.type in ('income', 'expense') else 'expense'))
    acct_id = get_or_create(conn, 'accounts', 'name', t.account) if t.account else None
    acct_to_id = get_or_create(conn, 'accounts', 'name', t.account_to) if t.account_to else None

    if t.type == 'transfer':
        if not acct_id or not acct_to_id:
            conn.close()
            raise HTTPException(400, "Transfer butuh akun asal dan akun tujuan")
        if acct_id == acct_to_id:
            conn.close()
            raise HTTPException(400, "Akun asal dan tujuan tidak boleh sama")

    conn.execute(
        "INSERT INTO transactions(date, type, amount, category_id, account_id, account_to_id, note, merchant) "
        "VALUES (?,?,?,?,?,?,?,?)",
        (tx_date, t.type, t.amount, cat_id, acct_id, acct_to_id, t.note, t.merchant)
    )
    tx_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    tx = conn.execute("SELECT * FROM transactions WHERE id=?", (tx_id,)).fetchone()
    _apply_balance(conn, tx, +1)
    if cat_id and t.type == 'expense':
        _sync_budget_spent(conn, tx_date[:7], cat_id)
    conn.commit()
    row = conn.execute(TX_SELECT + " WHERE t.id=?", (tx_id,)).fetchone()
    conn.close()
    return {"status": "ok", "transaction": dict(row)}


@router.patch("/transactions/{tx_id}")
def update_transaction(tx_id: int, p: TransactionPatch):
    """Ubah transaksi. Efek saldo lama dibatalkan lalu efek baru diterapkan."""
    conn = get_db()
    old = conn.execute("SELECT * FROM transactions WHERE id=?", (tx_id,)).fetchone()
    if not old:
        conn.close()
        raise HTTPException(404, f"Transaksi {tx_id} tidak ditemukan")

    new_type = p.type if p.type is not None else old["type"]
    _validate_type(new_type)
    new_amount = p.amount if p.amount is not None else old["amount"]
    _validate_amount(new_amount)
    new_date = _validate_date(p.date) if p.date is not None else old["date"]

    cat_id = old["category_id"]
    if p.category is not None:
        cat_id = get_or_create(conn, 'categories', 'name', p.category) if p.category else None
    acct_id = old["account_id"]
    if p.account is not None:
        acct_id = get_or_create(conn, 'accounts', 'name', p.account) if p.account else None
    acct_to_id = old["account_to_id"]
    if p.account_to is not None:
        acct_to_id = get_or_create(conn, 'accounts', 'name', p.account_to) if p.account_to else None

    if new_type == 'transfer':
        if not acct_id or not acct_to_id:
            conn.close()
            raise HTTPException(400, "Transfer butuh akun asal dan akun tujuan")
        if acct_id == acct_to_id:
            conn.close()
            raise HTTPException(400, "Akun asal dan tujuan tidak boleh sama")

    _apply_balance(conn, old, -1)          # batalkan efek lama
    conn.execute(
        "UPDATE transactions SET date=?, type=?, amount=?, category_id=?, account_id=?, "
        "account_to_id=?, note=?, merchant=? WHERE id=?",
        (new_date, new_type, new_amount, cat_id, acct_id, acct_to_id,
         p.note if p.note is not None else old["note"],
         p.merchant if p.merchant is not None else old["merchant"],
         tx_id)
    )
    new = conn.execute("SELECT * FROM transactions WHERE id=?", (tx_id,)).fetchone()
    _apply_balance(conn, new, +1)          # terapkan efek baru

    for month in {old["date"][:7], new_date[:7]}:
        _sync_budget_spent(conn, month)
    conn.commit()
    row = conn.execute(TX_SELECT + " WHERE t.id=?", (tx_id,)).fetchone()
    conn.close()
    return {"status": "ok", "transaction": dict(row)}


@router.delete("/transactions/{tx_id}")
def delete_transaction(tx_id: int):
    conn = get_db()
    tx = conn.execute("SELECT * FROM transactions WHERE id=?", (tx_id,)).fetchone()
    if not tx:
        conn.close()
        raise HTTPException(404, f"Transaksi {tx_id} tidak ditemukan")
    _apply_balance(conn, tx, -1)
    month = tx["date"][:7]
    inst_id = tx["installment_id"] if "installment_id" in tx.keys() else None
    conn.execute("DELETE FROM transactions WHERE id=?", (tx_id,))
    # cicilan yang pembayarannya dihapus harus ikut turun hitungannya
    if inst_id:
        conn.execute(
            "UPDATE installments SET paid_installments = MAX(0, paid_installments - 1), active = 1 "
            "WHERE id=?", (inst_id,)
        )
    _sync_budget_spent(conn, month)
    conn.commit()
    conn.close()
    return {"status": "ok", "deleted": tx_id}


@router.get("/transactions")
def list_transactions(
    limit: int = Query(20, ge=1, le=500),
    offset: int = Query(0, ge=0),
    month: Optional[str] = None,
    type: Optional[str] = None,
    category: Optional[str] = None,
    account: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    q: Optional[str] = None,
):
    """Daftar transaksi dengan filter opsional.

    `total` adalah jumlah baris yang cocok dengan filter (bukan jumlah yang
    dikembalikan), supaya paginasi di UI bisa benar.
    """
    where, params = [], []
    if month:
        where.append("t.date LIKE ?")
        params.append(_validate_month(month) + '%')
    if type:
        _validate_type(type)
        where.append("t.type = ?")
        params.append(type)
    if category:
        where.append("c.name = ?")
        params.append(category)
    if account:
        where.append("(a.name = ? OR a2.name = ?)")
        params += [account, account]
    if date_from:
        where.append("t.date >= ?")
        params.append(_validate_date(date_from))
    if date_to:
        where.append("t.date <= ?")
        params.append(_validate_date(date_to))
    if q:
        where.append("(COALESCE(t.note,'') LIKE ? OR COALESCE(t.merchant,'') LIKE ?)")
        params += [f"%{q}%", f"%{q}%"]

    clause = (" WHERE " + " AND ".join(where)) if where else ""
    conn = get_db()
    rows = conn.execute(
        TX_SELECT + clause + " ORDER BY t.date DESC, t.time DESC, t.id DESC LIMIT ? OFFSET ?",
        params + [limit, offset]
    ).fetchall()
    total = conn.execute(
        "SELECT COUNT(*) FROM transactions t "
        "LEFT JOIN categories c ON t.category_id=c.id "
        "LEFT JOIN accounts a ON t.account_id=a.id "
        "LEFT JOIN accounts a2 ON t.account_to_id=a2.id" + clause,
        params
    ).fetchone()[0]
    sums = conn.execute(
        "SELECT t.type, COALESCE(SUM(t.amount),0) AS total FROM transactions t "
        "LEFT JOIN categories c ON t.category_id=c.id "
        "LEFT JOIN accounts a ON t.account_id=a.id "
        "LEFT JOIN accounts a2 ON t.account_to_id=a2.id" + clause + " GROUP BY t.type",
        params
    ).fetchall()
    conn.close()
    return {
        "transactions": [dict(r) for r in rows],
        "total": total,
        "limit": limit,
        "offset": offset,
        "totals_by_type": {r["type"]: r["total"] for r in sums},
    }


# ── Summary ───────────────────────────────────────

@router.get("/summary")
def monthly_summary(month: Optional[str] = None):
    """Ringkasan bulanan: pemasukan, pengeluaran, tabungan, sisa."""
    month = _validate_month(month) if month else date.today().isoformat()[:7]
    conn = get_db()

    by_type = {r["type"]: r["total"] for r in conn.execute(
        "SELECT type, COALESCE(SUM(amount),0) AS total FROM transactions "
        "WHERE date LIKE ? GROUP BY type", (month + '%',)
    ).fetchall()}
    income = by_type.get('income', 0)
    expense = by_type.get('expense', 0)
    saving = by_type.get('saving', 0)
    invest = by_type.get('invest', 0)

    cat_spending = conn.execute("""
        SELECT c.name, c.icon, SUM(t.amount) as total
        FROM transactions t JOIN categories c ON t.category_id = c.id
        WHERE t.type='expense' AND t.date LIKE ?
        GROUP BY c.id ORDER BY total DESC
    """, (month + '%',)).fetchall()

    _sync_budget_spent(conn, month)
    conn.commit()
    budgets = conn.execute("""
        SELECT b.*, c.name, c.icon FROM budgets b
        JOIN categories c ON b.category_id = c.id
        WHERE b.month=? ORDER BY c.name
    """, (month,)).fetchall()

    accounts = conn.execute("SELECT * FROM accounts ORDER BY name").fetchall()
    installments = conn.execute(
        "SELECT * FROM installments WHERE active=1 ORDER BY start_date"
    ).fetchall()
    today = date.today().isoformat()
    due = conn.execute(
        "SELECT COUNT(*) FROM recurring WHERE active=1 AND next_date IS NOT NULL AND next_date<=?",
        (today,)
    ).fetchone()[0]
    conn.close()

    return {
        "month": month,
        "income": income,
        "expense": expense,
        "saving": saving,
        "invest": invest,
        "balance": income - expense,
        "net_worth": sum(a["balance"] for a in accounts),
        "recurring_due": due,
        "cat_spending": [dict(r) for r in cat_spending],
        "budgets": [dict(r) for r in budgets],
        "accounts": [dict(r) for r in accounts],
        "installments": [dict(r) for r in installments],
    }


@router.get("/trend")
def monthly_trend(months: int = Query(6, ge=1, le=36)):
    """Pemasukan vs pengeluaran beberapa bulan terakhir, untuk grafik tren."""
    today = date.today()
    labels = []
    for i in range(months - 1, -1, -1):
        d = _month_shift(date(today.year, today.month, 1), -i, 1)
        labels.append(d.isoformat()[:7])

    conn = get_db()
    rows = conn.execute(
        "SELECT substr(date,1,7) AS m, type, COALESCE(SUM(amount),0) AS total "
        "FROM transactions WHERE substr(date,1,7) >= ? GROUP BY m, type",
        (labels[0],)
    ).fetchall()
    conn.close()

    agg = {m: {"income": 0, "expense": 0, "saving": 0, "invest": 0} for m in labels}
    for r in rows:
        if r["m"] in agg and r["type"] in agg[r["m"]]:
            agg[r["m"]][r["type"]] = r["total"]
    return {"months": labels, "series": [{"month": m, **agg[m]} for m in labels]}


# ── Budgets ───────────────────────────────────────

@router.get("/budgets")
def list_budgets(month: Optional[str] = None):
    """Budget beserta realisasi. `spent` selalu dihitung dari transaksi."""
    month = _validate_month(month) if month else date.today().isoformat()[:7]
    conn = get_db()
    _sync_budget_spent(conn, month)
    conn.commit()
    rows = conn.execute("""
        SELECT b.*, c.name as cat_name, c.icon FROM budgets b
        JOIN categories c ON b.category_id = c.id
        WHERE b.month=? ORDER BY c.name
    """, (month,)).fetchall()
    conn.close()
    out = []
    for r in rows:
        d = dict(r)
        d["remaining"] = round((d["amount"] or 0) - (d["spent"] or 0), 2)
        d["pct"] = round((d["spent"] or 0) / d["amount"] * 100, 1) if d["amount"] else 0
        d["over"] = d["remaining"] < 0
        out.append(d)
    return {"month": month, "budgets": out}


@router.post("/budgets")
def set_budget(b: BudgetIn):
    _validate_amount(b.amount)
    month = _validate_month(b.month)
    conn = get_db()
    cat_id = get_or_create(conn, 'categories', 'name', b.category)
    existing = conn.execute(
        "SELECT id FROM budgets WHERE category_id=? AND month=?", (cat_id, month)
    ).fetchone()
    if existing:
        conn.execute("UPDATE budgets SET amount=? WHERE id=?", (b.amount, existing['id']))
    else:
        conn.execute("INSERT INTO budgets(category_id, month, amount) VALUES (?,?,?)",
                     (cat_id, month, b.amount))
    _sync_budget_spent(conn, month, cat_id)
    conn.commit()
    conn.close()
    return {"status": "ok"}


@router.delete("/budgets/{budget_id}")
def delete_budget(budget_id: int):
    conn = get_db()
    row = conn.execute("SELECT id FROM budgets WHERE id=?", (budget_id,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(404, f"Budget {budget_id} tidak ditemukan")
    conn.execute("DELETE FROM budgets WHERE id=?", (budget_id,))
    conn.commit()
    conn.close()
    return {"status": "ok", "deleted": budget_id}


# ── Recurring ─────────────────────────────────────

@router.get("/recurring")
def list_recurring():
    conn = get_db()
    rows = conn.execute("""
        SELECT r.*, c.name as cat_name, c.icon, a.name as acct_name
        FROM recurring r
        LEFT JOIN categories c ON r.category_id = c.id
        LEFT JOIN accounts a ON r.account_id = a.id
        ORDER BY r.active DESC, r.day
    """).fetchall()
    conn.close()
    today = date.today().isoformat()
    out = []
    for r in rows:
        d = dict(r)
        d["due"] = bool(d["active"] and d["next_date"] and d["next_date"] <= today)
        out.append(d)
    return {"recurring": out, "due_count": sum(1 for x in out if x["due"])}


@router.post("/recurring")
def add_recurring(r: RecurringIn):
    _validate_type(r.type)
    _validate_amount(r.amount)
    if r.frequency not in FREQUENCIES:
        raise HTTPException(400, f"Frekuensi tidak valid: {r.frequency}. Pilihan: {', '.join(FREQUENCIES)}")

    conn = get_db()
    cat_id = get_or_create(conn, 'categories', 'name', r.category) if r.category else None
    acct_id = get_or_create(conn, 'accounts', 'name', r.account) if r.account else None
    next_date = _first_due(r.frequency, r.day).isoformat()
    conn.execute(
        "INSERT INTO recurring(name, type, amount, category_id, account_id, frequency, day, next_date) "
        "VALUES (?,?,?,?,?,?,?,?)",
        (r.name, r.type, r.amount, cat_id, acct_id, r.frequency, r.day, next_date)
    )
    rid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.commit()
    conn.close()
    return {"status": "ok", "id": rid, "next_date": next_date}


@router.patch("/recurring/{rec_id}")
def update_recurring(rec_id: int, p: RecurringPatch):
    conn = get_db()
    old = conn.execute("SELECT * FROM recurring WHERE id=?", (rec_id,)).fetchone()
    if not old:
        conn.close()
        raise HTTPException(404, f"Recurring {rec_id} tidak ditemukan")
    if p.amount is not None:
        _validate_amount(p.amount)
    if p.frequency is not None and p.frequency not in FREQUENCIES:
        conn.close()
        raise HTTPException(400, f"Frekuensi tidak valid: {p.frequency}")

    name = p.name if p.name is not None else old["name"]
    amount = p.amount if p.amount is not None else old["amount"]
    day = p.day if p.day is not None else old["day"]
    freq = p.frequency if p.frequency is not None else old["frequency"]
    active = int(p.active) if p.active is not None else old["active"]

    next_date = old["next_date"]
    if p.day is not None or p.frequency is not None or (p.active and not old["active"]):
        next_date = _first_due(freq, day).isoformat()

    conn.execute(
        "UPDATE recurring SET name=?, amount=?, day=?, frequency=?, active=?, next_date=? WHERE id=?",
        (name, amount, day, freq, active, next_date, rec_id)
    )
    conn.commit()
    row = conn.execute("SELECT * FROM recurring WHERE id=?", (rec_id,)).fetchone()
    conn.close()
    return {"status": "ok", "recurring": dict(row)}


@router.delete("/recurring/{rec_id}")
def delete_recurring(rec_id: int):
    conn = get_db()
    row = conn.execute("SELECT id FROM recurring WHERE id=?", (rec_id,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(404, f"Recurring {rec_id} tidak ditemukan")
    conn.execute("DELETE FROM recurring WHERE id=?", (rec_id,))
    conn.commit()
    conn.close()
    return {"status": "ok", "deleted": rec_id}


@router.post("/recurring/run-due")
def run_due_recurring(until: Optional[str] = None):
    """Buat transaksi untuk semua recurring yang sudah jatuh tempo.

    Idempoten: `next_date` dimajukan setelah setiap pembuatan, dan
    kombinasi (recurring_id, date) dicek agar tidak dobel walau endpoint
    dipanggil berulang kali.
    """
    limit_date = date.fromisoformat(_validate_date(until)) if until else date.today()
    conn = get_db()
    rows = conn.execute("SELECT * FROM recurring WHERE active=1 AND next_date IS NOT NULL").fetchall()
    created = []
    for r in rows:
        cursor = date.fromisoformat(r["next_date"])
        guard = 0
        while cursor <= limit_date and guard < 120:
            guard += 1
            exists = conn.execute(
                "SELECT id FROM transactions WHERE recurring_id=? AND date=?",
                (r["id"], cursor.isoformat())
            ).fetchone()
            if not exists:
                conn.execute(
                    "INSERT INTO transactions(date, type, amount, category_id, account_id, note, recurring_id) "
                    "VALUES (?,?,?,?,?,?,?)",
                    (cursor.isoformat(), r["type"], r["amount"], r["category_id"],
                     r["account_id"], f"[otomatis] {r['name']}", r["id"])
                )
                tx_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                tx = conn.execute("SELECT * FROM transactions WHERE id=?", (tx_id,)).fetchone()
                _apply_balance(conn, tx, +1)
                _sync_budget_spent(conn, cursor.isoformat()[:7], r["category_id"])
                created.append({"id": tx_id, "name": r["name"], "date": cursor.isoformat(),
                                "type": r["type"], "amount": r["amount"]})
            conn.execute("UPDATE recurring SET last_run=? WHERE id=?", (cursor.isoformat(), r["id"]))
            cursor = _advance(r["frequency"], r["day"], cursor)
        conn.execute("UPDATE recurring SET next_date=? WHERE id=?", (cursor.isoformat(), r["id"]))
    conn.commit()
    conn.close()
    return {"status": "ok", "created": created, "count": len(created)}


# ── Installments ──────────────────────────────────

@router.post("/installments")
def add_installment(inst: InstallmentIn):
    _validate_amount(inst.total_amount)
    _validate_amount(inst.monthly_amount)
    if inst.total_installments < 1:
        raise HTTPException(400, "Jumlah cicilan minimal 1")
    start = _validate_date(inst.start_date) or date.today().isoformat()

    conn = get_db()
    cat_id = get_or_create(conn, 'categories', 'name', inst.category) if inst.category else None
    acct_id = get_or_create(conn, 'accounts', 'name', inst.account) if inst.account else None
    end = _month_shift(date.fromisoformat(start), inst.total_installments - 1,
                       date.fromisoformat(start).day).isoformat()
    conn.execute(
        "INSERT INTO installments(name, total_amount, total_installments, monthly_amount, "
        "category_id, account_id, start_date, end_date) VALUES (?,?,?,?,?,?,?,?)",
        (inst.name, inst.total_amount, inst.total_installments, inst.monthly_amount,
         cat_id, acct_id, start, end)
    )
    iid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.commit()
    conn.close()
    return {"status": "ok", "id": iid, "end_date": end}


@router.get("/installments")
def list_installments():
    conn = get_db()
    rows = conn.execute("""
        SELECT i.*, c.name as cat_name, c.icon, a.name as acct_name
        FROM installments i
        LEFT JOIN categories c ON i.category_id = c.id
        LEFT JOIN accounts a ON i.account_id = a.id
        ORDER BY i.active DESC, i.start_date
    """).fetchall()
    conn.close()
    out = []
    for r in rows:
        d = dict(r)
        left = max(0, (d["total_installments"] or 0) - (d["paid_installments"] or 0))
        d["remaining_count"] = left
        d["remaining_amount"] = round(left * (d["monthly_amount"] or 0), 2)
        d["progress_pct"] = round((d["paid_installments"] or 0) / d["total_installments"] * 100, 1) \
            if d["total_installments"] else 0
        out.append(d)
    return {"installments": out}


@router.post("/installments/{inst_id}/pay")
def pay_installment(inst_id: int, on_date: Optional[str] = None):
    """Catat satu pembayaran cicilan sebagai transaksi pengeluaran."""
    conn = get_db()
    inst = conn.execute("SELECT * FROM installments WHERE id=?", (inst_id,)).fetchone()
    if not inst:
        conn.close()
        raise HTTPException(404, f"Cicilan {inst_id} tidak ditemukan")
    if (inst["paid_installments"] or 0) >= inst["total_installments"]:
        conn.close()
        raise HTTPException(400, "Cicilan sudah lunas")

    pay_date = _validate_date(on_date) or date.today().isoformat()
    seq = (inst["paid_installments"] or 0) + 1
    conn.execute(
        "INSERT INTO transactions(date, type, amount, category_id, account_id, note, installment_id) "
        "VALUES (?,?,?,?,?,?,?)",
        (pay_date, 'expense', inst["monthly_amount"], inst["category_id"], inst["account_id"],
         f"Cicilan {inst['name']} ({seq}/{inst['total_installments']})", inst_id)
    )
    tx_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    tx = conn.execute("SELECT * FROM transactions WHERE id=?", (tx_id,)).fetchone()
    _apply_balance(conn, tx, +1)
    done = seq >= inst["total_installments"]
    conn.execute("UPDATE installments SET paid_installments=?, active=? WHERE id=?",
                 (seq, 0 if done else 1, inst_id))
    _sync_budget_spent(conn, pay_date[:7], inst["category_id"])
    conn.commit()
    conn.close()
    return {"status": "ok", "transaction_id": tx_id, "paid": seq,
            "total": inst["total_installments"], "completed": done}


@router.delete("/installments/{inst_id}")
def delete_installment(inst_id: int):
    conn = get_db()
    row = conn.execute("SELECT id FROM installments WHERE id=?", (inst_id,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(404, f"Cicilan {inst_id} tidak ditemukan")
    conn.execute("DELETE FROM installments WHERE id=?", (inst_id,))
    conn.commit()
    conn.close()
    return {"status": "ok", "deleted": inst_id}


# ── Investments ───────────────────────────────────

@router.post("/investments")
def add_investment(inv: InvestmentIn):
    _validate_amount(inv.amount)
    conn = get_db()
    acct_id = get_or_create(conn, 'accounts', 'name', inv.account) if inv.account else None
    try:
        conn.execute(
            "INSERT INTO investments(name, type, amount, quantity, buy_price, account_id, "
            "maturity_date, return_rate, note) VALUES (?,?,?,?,?,?,?,?,?)",
            (inv.name, inv.type, inv.amount, inv.quantity, inv.buy_price, acct_id,
             inv.maturity_date, inv.return_rate, inv.note)
        )
        conn.commit()
    except sqlite3.IntegrityError as e:
        conn.close()
        raise HTTPException(400, f"Tipe investasi tidak valid: {inv.type}") from e
    iid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.close()
    return {"status": "ok", "id": iid}


@router.delete("/investments/{inv_id}")
def delete_investment(inv_id: int):
    conn = get_db()
    row = conn.execute("SELECT id FROM investments WHERE id=?", (inv_id,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(404, f"Investasi {inv_id} tidak ditemukan")
    conn.execute("DELETE FROM investments WHERE id=?", (inv_id,))
    conn.commit()
    conn.close()
    return {"status": "ok", "deleted": inv_id}


@router.get("/investments")
def list_investments():
    conn = get_db()
    rows = conn.execute("""
        SELECT inv.*, a.name as acct_name
        FROM investments inv
        LEFT JOIN accounts a ON inv.account_id = a.id
        ORDER BY inv.created_at DESC
    """).fetchall()
    conn.close()

    investments = [dict(r) for r in rows]
    for inv in investments:
        if inv['type'] == 'crypto' and inv['name'].lower() in ('bitcoin', 'btc'):
            price = fetch_crypto_price('bitcoin')
            if price:
                inv['current_price'] = price
        elif inv['type'] == 'crypto' and inv['name'].lower() in ('ethereum', 'eth'):
            price = fetch_crypto_price('ethereum')
            if price:
                inv['current_price'] = price
        elif inv['type'] == 'gold':
            price = fetch_gold_price()
            if price:
                inv['current_price'] = price
    return {"investments": investments}


@router.get("/portfolio")
def portfolio_summary():
    """Total portfolio value: crypto + gold + deposits + other investments."""
    conn = get_db()
    rows = conn.execute("SELECT * FROM investments").fetchall()
    conn.close()

    total_cost = 0
    total_current = 0
    details = []
    for inv in rows:
        inv = dict(inv)
        cost = inv['amount']
        current = inv['amount']
        if inv['type'] == 'crypto' and inv['name'].lower() in ('bitcoin', 'btc'):
            price = fetch_crypto_price('bitcoin')
            if price and inv['quantity']:
                current = price * inv['quantity']
        elif inv['type'] == 'gold':
            price = fetch_gold_price()
            if price and inv['quantity']:
                current = price * inv['quantity']
        elif inv['type'] == 'deposit' and inv['maturity_date']:
            try:
                start = datetime.strptime(inv.get('created_at', '')[:10], '%Y-%m-%d') \
                    if inv.get('created_at') else None
                if start and inv['return_rate']:
                    days = (datetime.now() - start).days
                    current = inv['amount'] + inv['amount'] * (inv['return_rate'] / 100) * days / 365
            except Exception:
                pass

        total_cost += cost
        total_current += current
        pnl = current - cost
        details.append({
            "id": inv['id'], "name": inv['name'], "type": inv['type'],
            "cost": round(cost, 2), "current": round(current, 2),
            "pnl": round(pnl, 2), "pnl_pct": round((pnl / cost * 100) if cost else 0, 2)
        })

    return {
        "total_cost": round(total_cost, 2),
        "total_current": round(total_current, 2),
        "total_pnl": round(total_current - total_cost, 2),
        "total_pnl_pct": round((total_current - total_cost) / total_cost * 100, 2) if total_cost else 0,
        "details": details
    }


# ── Export ────────────────────────────────────────

@router.get("/export")
def export_excel(month: Optional[str] = None):
    """Export transaksi ke file Excel (.xlsx)."""
    month = _validate_month(month) if month else date.today().isoformat()[:7]

    conn = get_db()
    rows = conn.execute("""
        SELECT t.date, t.time, t.type, t.amount, c.name as cat, c.icon, a.name as acct,
               t.merchant, t.note
        FROM transactions t
        LEFT JOIN categories c ON t.category_id = c.id
        LEFT JOIN accounts a ON t.account_id = a.id
        WHERE t.date LIKE ?
        ORDER BY t.date DESC, t.time DESC
    """, (month + '%',)).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Transaksi {month}"

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="1e293b")
    headers = ["Tanggal", "Jam", "Tipe", "Jumlah", "Kategori", "Akun", "Merchant", "Catatan"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    type_colors = {"income": "22c55e", "expense": "ef4444", "saving": "00d4ff",
                   "invest": "a855f7", "transfer": "f59e0b"}
    for i, r in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=r['date'])
        ws.cell(row=i, column=2, value=r['time'])
        ws.cell(row=i, column=3, value=r['type'])
        amt = ws.cell(row=i, column=4, value=r['amount'])
        amt.number_format = '#,##0'
        amt.font = Font(color=type_colors.get(r['type'], '64748b'), bold=True)
        ws.cell(row=i, column=5, value=(r['icon'] or '') + ' ' + (r['cat'] or ''))
        ws.cell(row=i, column=6, value=r['acct'] or '')
        ws.cell(row=i, column=7, value=r['merchant'] or '')
        ws.cell(row=i, column=8, value=r['note'] or '')

    for col in range(1, 9):
        max_len = max((len(str(ws.cell(row=r, column=col).value or ''))
                       for r in range(1, len(rows) + 2)), default=10)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_len + 4, 30)

    summary_row = len(rows) + 3
    summary = {}
    for r in rows:
        summary[r['type']] = summary.get(r['type'], 0) + r['amount']
    ws.cell(row=summary_row, column=1, value="RINGKASAN").font = Font(bold=True, size=12)
    row = summary_row + 1
    for typ, total in summary.items():
        ws.cell(row=row, column=1, value=f"{typ}:")
        c = ws.cell(row=row, column=2, value=total)
        c.number_format = '#,##0'
        c.font = Font(bold=True, color=type_colors.get(typ, 'ffffff'))
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=finance_{month}.xlsx"}
    )
